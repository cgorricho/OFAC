"""Unit tests for report generator service."""

import io

import pytest
from openpyxl import load_workbook

from ofac.core.models import (
    BatchScreeningResponse,
    EntityInput,
    MatchResult,
    MatchStatus,
    MatchType,
    OFACList,
    ScreeningResult,
)
from ofac.core.reporter import ReportGenerator


@pytest.fixture
def sample_batch_response() -> BatchScreeningResponse:
    """Create sample batch response for testing."""
    results = [
        ScreeningResult(
            entity_input=EntityInput(entity_name="Test Org 1", country="US"),
            match_status=MatchStatus.OK,
            highest_score=0,
            ofac_version="2025-12-01",
            matches=[],
        ),
        ScreeningResult(
            entity_input=EntityInput(entity_name="Test Org 2", country="SY"),
            match_status=MatchStatus.REVIEW,
            highest_score=85,
            ofac_version="2025-12-01",
            matches=[
                MatchResult(
                    sdn_name="TEST ENTITY",
                    sdn_type="entity",
                    match_score=85,
                    match_type=MatchType.FUZZY,
                    ofac_list=OFACList.SDN,
                    programs=[],
                    ent_num=1,
                    country="Syria",
                    country_match=True,
                )
            ],
        ),
    ]
    return BatchScreeningResponse(
        results=results,
        total_screened=2,
        ok_count=1,
        review_count=1,
        nok_count=0,
    )


class TestReportGenerator:
    """Tests for ReportGenerator class."""

    def test_generate_creates_workbook(self, sample_batch_response) -> None:
        """generate() creates a valid Excel workbook."""
        generator = ReportGenerator()
        workbook_bytes = generator.generate(sample_batch_response)

        # Verify it's a valid Excel file
        workbook = load_workbook(io.BytesIO(workbook_bytes))
        assert workbook is not None

    def test_generate_returns_bytes(self, sample_batch_response) -> None:
        """generate() returns bytes suitable for download."""
        generator = ReportGenerator()
        workbook_bytes = generator.generate(sample_batch_response)

        assert isinstance(workbook_bytes, bytes)
        assert len(workbook_bytes) > 0

    def test_generate_empty_results_raises_error(self) -> None:
        """generate() raises ValueError for empty results."""
        generator = ReportGenerator()
        empty_response = BatchScreeningResponse(
            results=[],
            total_screened=0,
            ok_count=0,
            review_count=0,
            nok_count=0,
        )

        with pytest.raises(ValueError, match="Cannot generate report from empty results"):
            generator.generate(empty_response)

    def test_generate_workbook_structure(self, sample_batch_response) -> None:
        """generate() creates workbook with proper structure."""
        generator = ReportGenerator()
        workbook_bytes = generator.generate(sample_batch_response)

        workbook = load_workbook(io.BytesIO(workbook_bytes))
        # Workbook should exist with Summary sheet
        assert workbook is not None
        assert "Summary" in workbook.sheetnames

    def test_summary_sheet_contains_metadata(self, sample_batch_response) -> None:
        """Summary sheet contains screening metadata."""
        generator = ReportGenerator()
        workbook_bytes = generator.generate(sample_batch_response)

        workbook = load_workbook(io.BytesIO(workbook_bytes))
        sheet = workbook["Summary"]

        # Check for key metadata
        cell_values = [cell.value for row in sheet.iter_rows() for cell in row]
        assert "OFAC Screening Report - Summary" in cell_values
        assert "Screening ID" in cell_values
        assert "Total Entities Screened" in cell_values
        assert "Status Breakdown" in cell_values

    def test_summary_sheet_contains_statistics(self, sample_batch_response) -> None:
        """Summary sheet contains status breakdown."""
        generator = ReportGenerator()
        workbook_bytes = generator.generate(sample_batch_response)

        workbook = load_workbook(io.BytesIO(workbook_bytes))
        sheet = workbook["Summary"]

        # Check for status breakdown
        cell_values = [cell.value for row in sheet.iter_rows() for cell in row]
        assert "OK" in cell_values
        assert "REVIEW" in cell_values
        assert "NOK" in cell_values or 0 in cell_values  # NOK count might be 0

    def test_details_sheet_exists(self, sample_batch_response) -> None:
        """Details sheet is created in workbook."""
        generator = ReportGenerator()
        workbook_bytes = generator.generate(sample_batch_response)

        workbook = load_workbook(io.BytesIO(workbook_bytes))
        assert "All Results" in workbook.sheetnames

    def test_details_sheet_contains_headers(self, sample_batch_response) -> None:
        """Details sheet contains expected column headers."""
        generator = ReportGenerator()
        workbook_bytes = generator.generate(sample_batch_response)

        workbook = load_workbook(io.BytesIO(workbook_bytes))
        sheet = workbook["All Results"]

        # Check headers (including audit trail fields)
        headers = [cell.value for cell in sheet[1]]
        assert "Row #" in headers
        assert "Organization Name" in headers
        assert "Status" in headers
        assert "Match Score" in headers
        assert "Screening ID" in headers
        assert "Timestamp" in headers
        assert "OFAC Version" in headers
        assert "SDN Entity ID" in headers
        assert "Programs" in headers
        assert "General License" in headers

    def test_details_sheet_contains_data(self, sample_batch_response) -> None:
        """Details sheet contains screening result data."""
        generator = ReportGenerator()
        workbook_bytes = generator.generate(sample_batch_response)

        workbook = load_workbook(io.BytesIO(workbook_bytes))
        sheet = workbook["All Results"]

        # Should have header row + 2 data rows
        assert sheet.max_row == 3
        # Check that data rows contain organization names
        cell_values = [cell.value for row in sheet.iter_rows(min_row=2) for cell in row]
        assert "Test Org 1" in cell_values
        assert "Test Org 2" in cell_values

    def test_exceptions_sheet_exists(self, sample_batch_response) -> None:
        """Exceptions sheet is created in workbook."""
        generator = ReportGenerator()
        workbook_bytes = generator.generate(sample_batch_response)

        workbook = load_workbook(io.BytesIO(workbook_bytes))
        assert "Exceptions" in workbook.sheetnames

    def test_exceptions_sheet_filters_review_nok(self) -> None:
        """Exceptions sheet contains only REVIEW and NOK cases."""
        results = [
            ScreeningResult(
                entity_input=EntityInput(entity_name="OK Org", country="US"),
                match_status=MatchStatus.OK,
                highest_score=0,
                ofac_version="2025-12-01",
                matches=[],
            ),
            ScreeningResult(
                entity_input=EntityInput(entity_name="REVIEW Org", country="SY"),
                match_status=MatchStatus.REVIEW,
                highest_score=85,
                ofac_version="2025-12-01",
                matches=[
                    MatchResult(
                        sdn_name="REVIEW ENTITY",
                        sdn_type="entity",
                        match_score=85,
                        match_type=MatchType.FUZZY,
                        ofac_list=OFACList.SDN,
                        programs=[],
                        ent_num=1,
                        country="Syria",
                        country_match=True,
                    )
                ],
            ),
            ScreeningResult(
                entity_input=EntityInput(entity_name="NOK Org", country="IR"),
                match_status=MatchStatus.NOK,
                highest_score=95,
                ofac_version="2025-12-01",
                matches=[
                    MatchResult(
                        sdn_name="NOK ENTITY",
                        sdn_type="entity",
                        match_score=95,
                        match_type=MatchType.EXACT,
                        ofac_list=OFACList.SDN,
                        programs=["IRAN"],
                        ent_num=2,
                        country="Iran",
                        country_match=True,
                    )
                ],
            ),
        ]
        batch_response = BatchScreeningResponse(
            results=results,
            total_screened=3,
            ok_count=1,
            review_count=1,
            nok_count=1,
        )

        generator = ReportGenerator()
        workbook_bytes = generator.generate(batch_response)

        workbook = load_workbook(io.BytesIO(workbook_bytes))
        sheet = workbook["Exceptions"]

        # Should have 2 rows (REVIEW + NOK), not OK
        cell_values = [cell.value for row in sheet.iter_rows() for cell in row]
        assert "REVIEW Org" in cell_values
        assert "NOK Org" in cell_values
        assert "OK Org" not in cell_values

    def test_exceptions_sheet_sorted_by_risk(self) -> None:
        """Exceptions sheet is sorted by status (NOK first) then score."""
        results = [
            ScreeningResult(
                entity_input=EntityInput(entity_name="REVIEW Low", country="SY"),
                match_status=MatchStatus.REVIEW,
                highest_score=80,
                ofac_version="2025-12-01",
                matches=[
                    MatchResult(
                        sdn_name="ENTITY",
                        sdn_type="entity",
                        match_score=80,
                        match_type=MatchType.FUZZY,
                        ofac_list=OFACList.SDN,
                        programs=[],
                        ent_num=1,
                        country="Syria",
                    )
                ],
            ),
            ScreeningResult(
                entity_input=EntityInput(entity_name="NOK High", country="IR"),
                match_status=MatchStatus.NOK,
                highest_score=98,
                ofac_version="2025-12-01",
                matches=[
                    MatchResult(
                        sdn_name="NOK ENTITY",
                        sdn_type="entity",
                        match_score=98,
                        match_type=MatchType.EXACT,
                        ofac_list=OFACList.SDN,
                        programs=["IRAN"],
                        ent_num=2,
                        country="Iran",
                    )
                ],
            ),
            ScreeningResult(
                entity_input=EntityInput(entity_name="REVIEW High", country="SY"),
                match_status=MatchStatus.REVIEW,
                highest_score=90,
                ofac_version="2025-12-01",
                matches=[
                    MatchResult(
                        sdn_name="REVIEW ENTITY",
                        sdn_type="entity",
                        match_score=90,
                        match_type=MatchType.FUZZY,
                        ofac_list=OFACList.SDN,
                        programs=[],
                        ent_num=3,
                        country="Syria",
                    )
                ],
            ),
        ]
        batch_response = BatchScreeningResponse(
            results=results,
            total_screened=3,
            ok_count=0,
            review_count=2,
            nok_count=1,
        )

        generator = ReportGenerator()
        workbook_bytes = generator.generate(batch_response)

        workbook = load_workbook(io.BytesIO(workbook_bytes))
        sheet = workbook["Exceptions"]

        # First data row should be NOK (highest risk)
        data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        assert data_rows[0][3] == "NOK"  # Status column
        assert data_rows[0][1] == "NOK High"  # Organization name

    def test_color_coding_applied(self) -> None:
        """Color coding is applied to status cells."""
        results = [
            ScreeningResult(
                entity_input=EntityInput(entity_name="OK Org", country="US"),
                match_status=MatchStatus.OK,
                highest_score=0,
                ofac_version="2025-12-01",
                matches=[],
            ),
            ScreeningResult(
                entity_input=EntityInput(entity_name="REVIEW Org", country="SY"),
                match_status=MatchStatus.REVIEW,
                highest_score=85,
                ofac_version="2025-12-01",
                matches=[
                    MatchResult(
                        sdn_name="ENTITY",
                        sdn_type="entity",
                        match_score=85,
                        match_type=MatchType.FUZZY,
                        ofac_list=OFACList.SDN,
                        programs=[],
                        ent_num=1,
                        country="Syria",
                    )
                ],
            ),
            ScreeningResult(
                entity_input=EntityInput(entity_name="NOK Org", country="IR"),
                match_status=MatchStatus.NOK,
                highest_score=95,
                ofac_version="2025-12-01",
                matches=[
                    MatchResult(
                        sdn_name="NOK ENTITY",
                        sdn_type="entity",
                        match_score=95,
                        match_type=MatchType.EXACT,
                        ofac_list=OFACList.SDN,
                        programs=["IRAN"],
                        ent_num=2,
                        country="Iran",
                    )
                ],
            ),
        ]
        batch_response = BatchScreeningResponse(
            results=results,
            total_screened=3,
            ok_count=1,
            review_count=1,
            nok_count=1,
        )

        generator = ReportGenerator()
        workbook_bytes = generator.generate(batch_response)

        workbook = load_workbook(io.BytesIO(workbook_bytes))
        sheet = workbook["All Results"]

        # Check color coding (status column is D, row 2 is first data row)
        ok_cell = sheet["D2"]  # OK status
        review_cell = sheet["D3"]  # REVIEW status
        nok_cell = sheet["D4"]  # NOK status

        # openpyxl uses "00" prefix for RGB (not "FF")
        assert ok_cell.fill.start_color.rgb == "00C6F6D5"  # Light green
        assert review_cell.fill.start_color.rgb == "00FEFCBF"  # Light yellow
        assert nok_cell.fill.start_color.rgb == "00FED7D7"  # Light red
        assert nok_cell.font.bold is True  # NOK is bolded

