"""Report generation service for OFAC screening results.

This module provides the ReportGenerator class for creating Excel reports
from screening results with multiple sheets, color coding, and audit trail fields.

Usage:
    from ofac.core.reporter import ReportGenerator
    from ofac.core.models import BatchScreeningResponse

    generator = ReportGenerator()
    workbook_bytes = generator.generate(batch_response)
"""

import io
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ofac.core.models import MatchStatus

# Color coding per UX spec
COLOR_OK = "C6F6D5"  # Light green
COLOR_REVIEW = "FEFCBF"  # Light yellow
COLOR_NOK = "FED7D7"  # Light red

if TYPE_CHECKING:
    from ofac.core.models import BatchScreeningResponse


class ReportGenerator:
    """Generates Excel reports from OFAC screening results.

    Creates multi-sheet Excel workbooks with:
    - Summary sheet with statistics
    - Detailed results sheet
    - Exceptions sheet (REVIEW/NOK only)
    - Color coding and professional formatting
    - Complete audit trail fields

    Attributes:
        workbook: The openpyxl Workbook instance

    Example:
        generator = ReportGenerator()
        workbook_bytes = generator.generate(batch_response)
        with open("report.xlsx", "wb") as f:
            f.write(workbook_bytes)
    """

    def __init__(self) -> None:
        """Initialize the report generator."""
        self.workbook: Workbook | None = None

    def generate(self, batch_response: "BatchScreeningResponse") -> bytes:
        """Generate Excel report from batch screening results.

        Args:
            batch_response: BatchScreeningResponse containing all screening results.

        Returns:
            bytes: Excel workbook as bytes for download.

        Raises:
            ValueError: If batch_response is empty or invalid.
        """
        if not batch_response.results:
            raise ValueError("Cannot generate report from empty results")

        self.workbook = Workbook()
        # Remove default sheet, we'll create our own
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet)

        # Generate sheets
        self._create_summary_sheet(batch_response)
        self._create_details_sheet(batch_response)
        self._create_exceptions_sheet(batch_response)

        # Save to bytes
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output.read()

    def _create_summary_sheet(self, batch_response: "BatchScreeningResponse") -> None:
        """Create Summary sheet with statistics.

        Args:
            batch_response: BatchScreeningResponse containing screening results.
        """
        sheet = self.workbook.create_sheet("Summary")

        # Header
        sheet.append(["OFAC Screening Report - Summary"])
        sheet.append([])  # Empty row

        # Screening metadata
        sheet.append(["Screening Information"])
        if batch_response.results:
            first_result = batch_response.results[0]
            sheet.append(["Screening ID", first_result.screening_id])
            sheet.append(
                ["Screening Date", first_result.timestamp.strftime("%Y-%m-%d")]
            )
            sheet.append(
                ["Screening Time", first_result.timestamp.strftime("%H:%M:%S UTC")]
            )
            sheet.append(["OFAC Data Version", first_result.ofac_version or "Unknown"])
        else:
            sheet.append(["Screening ID", "N/A"])
            sheet.append(["Screening Date", "N/A"])
            sheet.append(["Screening Time", "N/A"])
            sheet.append(
                ["OFAC Data Version", batch_response.ofac_version or "Unknown"]
            )

        sheet.append([])  # Empty row

        # Statistics
        sheet.append(["Screening Statistics"])
        sheet.append(["Total Entities Screened", batch_response.total_screened])
        sheet.append([])  # Empty row

        # Status breakdown
        sheet.append(["Status Breakdown"])
        sheet.append(["Status", "Count", "Percentage"])

        total = batch_response.total_screened
        if total > 0:
            ok_pct = (batch_response.ok_count / total) * 100
            review_pct = (batch_response.review_count / total) * 100
            nok_pct = (batch_response.nok_count / total) * 100
        else:
            ok_pct = review_pct = nok_pct = 0.0

        sheet.append(["OK", batch_response.ok_count, f"{ok_pct:.1f}%"])
        sheet.append(["REVIEW", batch_response.review_count, f"{review_pct:.1f}%"])
        sheet.append(["NOK", batch_response.nok_count, f"{nok_pct:.1f}%"])

        # Auto-adjust column widths
        sheet.column_dimensions["A"].width = 25
        sheet.column_dimensions["B"].width = 15
        sheet.column_dimensions["C"].width = 15

    def _create_details_sheet(self, batch_response: "BatchScreeningResponse") -> None:
        """Create Detailed Results sheet with all screening results.

        Args:
            batch_response: BatchScreeningResponse containing screening results.
        """
        sheet = self.workbook.create_sheet("All Results")

        # Header row
        headers = [
            "Row #",
            "Organization Name",
            "Country",
            "Status",
            "Match Score",
            "Matched SDN",
            "Match Type",
            "Country Alignment",
        ]
        sheet.append(headers)

        # Data rows
        for idx, result in enumerate(batch_response.results, start=1):
            entity_input = result.entity_input
            highest_match = result.matches[0] if result.matches else None

            row = [
                idx,
                entity_input.entity_name,
                entity_input.country or "N/A",
                result.match_status.value,
                result.highest_score,
                highest_match.sdn_name if highest_match else "N/A",
                highest_match.match_type.value if highest_match else "N/A",
                "Match"
                if highest_match and highest_match.country_match
                else "N/A"
                if not highest_match
                else "Mismatch",
            ]
            sheet.append(row)

        # Sort by status (NOK first, then REVIEW, then OK)
        # Note: openpyxl doesn't support sorting directly, so we'll sort the data before adding
        # For now, we'll add sorting in a future enhancement if needed

        # Apply color coding to status column
        self._apply_color_coding(sheet, status_col_idx=4, start_row=2)

        # Auto-adjust column widths
        for col_idx, _header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            if col_idx == 1:  # Row #
                sheet.column_dimensions[col_letter].width = 8
            elif col_idx == 2:  # Organization Name
                sheet.column_dimensions[col_letter].width = 30
            elif col_idx == 3:  # Country
                sheet.column_dimensions[col_letter].width = 15
            elif col_idx == 4 or col_idx == 5:  # Status
                sheet.column_dimensions[col_letter].width = 12
            elif col_idx == 6:  # Matched SDN
                sheet.column_dimensions[col_letter].width = 35
            elif col_idx == 7:  # Match Type
                sheet.column_dimensions[col_letter].width = 12
            elif col_idx == 8:  # Country Alignment
                sheet.column_dimensions[col_letter].width = 18

    def _create_exceptions_sheet(
        self, batch_response: "BatchScreeningResponse"
    ) -> None:
        """Create Exceptions sheet with only REVIEW and NOK cases.

        Args:
            batch_response: BatchScreeningResponse containing screening results.
        """
        sheet = self.workbook.create_sheet("Exceptions")

        # Filter to only REVIEW and NOK
        exceptions = [
            r
            for r in batch_response.results
            if r.match_status in (MatchStatus.REVIEW, MatchStatus.NOK)
        ]

        if not exceptions:
            sheet.append(["No Exceptions"])
            sheet.append(["All entities screened as OK"])
            return

        # Sort by status (NOK first, then REVIEW), then by score descending
        exceptions.sort(
            key=lambda x: (
                0 if x.match_status == MatchStatus.NOK else 1,  # NOK first
                -x.highest_score,  # Higher scores first
            )
        )

        # Header row (same as details sheet)
        headers = [
            "Row #",
            "Organization Name",
            "Country",
            "Status",
            "Match Score",
            "Matched SDN",
            "Match Type",
            "Country Alignment",
        ]
        sheet.append(headers)

        # Data rows
        for idx, result in enumerate(exceptions, start=1):
            entity_input = result.entity_input
            highest_match = result.matches[0] if result.matches else None

            row = [
                idx,
                entity_input.entity_name,
                entity_input.country or "N/A",
                result.match_status.value,
                result.highest_score,
                highest_match.sdn_name if highest_match else "N/A",
                highest_match.match_type.value if highest_match else "N/A",
                "Match"
                if highest_match and highest_match.country_match
                else "N/A"
                if not highest_match
                else "Mismatch",
            ]
            sheet.append(row)

        # Auto-adjust column widths (same as details sheet)
        for col_idx, _header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            if col_idx == 1:  # Row #
                sheet.column_dimensions[col_letter].width = 8
            elif col_idx == 2:  # Organization Name
                sheet.column_dimensions[col_letter].width = 30
            elif col_idx == 3:  # Country
                sheet.column_dimensions[col_letter].width = 15
            elif col_idx == 4 or col_idx == 5:  # Status
                sheet.column_dimensions[col_letter].width = 12
            elif col_idx == 6:  # Matched SDN
                sheet.column_dimensions[col_letter].width = 35
            elif col_idx == 7:  # Match Type
                sheet.column_dimensions[col_letter].width = 12
            elif col_idx == 8:  # Country Alignment
                sheet.column_dimensions[col_letter].width = 18

        # Apply color coding to exceptions sheet
        if exceptions:  # Only if there are exceptions
            self._apply_color_coding(sheet, status_col_idx=4, start_row=2)

    def _apply_color_coding(
        self, sheet, status_col_idx: int, start_row: int = 2
    ) -> None:
        """Apply color coding to status cells based on match status.

        Args:
            sheet: The worksheet to apply formatting to.
            status_col_idx: Column index (1-based) containing status values.
            start_row: Row number to start applying formatting (default: 2, after header).
        """
        status_col_letter = get_column_letter(status_col_idx)

        for row_idx in range(start_row, sheet.max_row + 1):
            status_cell = sheet[f"{status_col_letter}{row_idx}"]
            status_value = status_cell.value

            if status_value == "OK":
                fill = PatternFill(
                    start_color=COLOR_OK, end_color=COLOR_OK, fill_type="solid"
                )
                status_cell.fill = fill
            elif status_value == "REVIEW":
                fill = PatternFill(
                    start_color=COLOR_REVIEW, end_color=COLOR_REVIEW, fill_type="solid"
                )
                status_cell.fill = fill
            elif status_value == "NOK":
                fill = PatternFill(
                    start_color=COLOR_NOK, end_color=COLOR_NOK, fill_type="solid"
                )
                status_cell.fill = fill
                # Also bold NOK text for accessibility
                status_cell.font = Font(bold=True)


__all__ = ["ReportGenerator"]
